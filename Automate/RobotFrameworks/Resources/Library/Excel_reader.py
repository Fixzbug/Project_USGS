import pandas  
import openpyxl  
import json
import warnings
import psutil
import pandas as pd
import time
import win32com.client as win32  # 306
from openpyxl import load_workbook
from pprint import pprint

from openpyxl.utils.exceptions import InvalidFileException
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

global  read_before_write
read_before_write = False


def Read_Excel_by_Sheet(filepath, sheet_name):

    # อ่านข้อมูลจาก Excel ด้วย pandas
    excel_data_df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)

    # ใช้เมธอด strip เพื่อลบช่องว่างด้านหน้าและด้านหลังในแต่ละเซลล์
    excel_data_df = excel_data_df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))

    # แปลง DataFrame ที่ทำความสะอาดแล้วเป็น JSON
    json_str = excel_data_df.to_json(orient='records')
    json_obj = json.loads(json_str)

    # นับจำนวนแถว
    row_count = len(excel_data_df.index)
    # นับจำนวนคอลัมน์
    column_count = len(excel_data_df.columns)

    # แสดงผลข้อมูล จำนวนแถว และจำนวนคอลัมน์
    # jsons = json.dumps(json_obj, ensure_ascii=False, sort_keys=True, indent=4, separators=(",", ": "))
    # print("Data:", jsons, "ROW:", row_count, "COL:", column_count)

    if json_obj:
        read_before_write = True
        return json_obj, row_count, column_count
    

def Write_Excel_by_Sheet(filepath, sheet_name, row_expect, column_expect, data):
    workbook = openpyxl.load_workbook(filename=filepath)
    sheet = workbook[sheet_name]
    sheet.cell(row=int(row_expect), column=int(column_expect)).value = data.strip()
    workbook.save(filepath)
    workbook.close()


def kill_excel_processes():
    for proc in psutil.process_iter(['pid', 'name']):
        if proc.info['name'] == 'EXCEL.EXE' or proc.info['name'].lower() == 'excel.exe':
            try:
                proc.terminate()
                proc.wait(timeout=3)
                time.sleep(1)
                if proc.is_running():
                    proc.kill()
                    print(f"Excel process with PID {proc.info['pid']} is still running!")
                else:
                    print(f"Terminated process {proc.info['name']} with PID {proc.info['pid']}")
            except (psutil.NoSuchProcess, psutil.AccessDenied, psutil.ZombieProcess) as e:
                print(f"Failed to terminate process {proc.info['name']} with PID {proc.info['pid']}: {e}")


def Write_Excel_by_SheetV1(filepath, sheet_name, row_expect, column_expect, data):

    # If reading before writing, kill Excel processes to ensure a clean state
    if  read_before_write:
        kill_excel_processes()

    # Initialize the Excel application
    excel = win32.Dispatch("Excel.Application")

    # Make Excel work in the background
    excel.Visible = False

    # Disable AutoRecover for this session
    excel.AutoRecover.Enabled = False

    # Open the specified workbook
    workbook = excel.Workbooks.Open(filepath)
    
    # Check if the specified sheet exists in the workbook
    try:
        sheet = workbook.Sheets(sheet_name)
    except Exception as e:
        print(f"Sheet '{sheet_name}' not found in the workbook.")
        workbook.Close(SaveChanges=False)
        excel.Quit()
        kill_excel_processes()
        return
    
    # Trim whitespace from the data
    trimmed_data = str(data).strip()

    # Write data to the specified cell
    sheet.Cells(int(row_expect), int(column_expect)).Value = trimmed_data

    # Close the workbook and Excel
    workbook.Close(SaveChanges=True)
    excel.Quit()

    # Ensure all Excel processes are terminated
    kill_excel_processes()

    # Print success message
    print("Data written successfully to", filepath, "in sheet", sheet_name, "starting at row", row_expect, "and column", column_expect, "data", data)
    # logging.info("Data written successfully to", filepath, "in sheet", sheet_name, "starting at row", row_expect, "and column", column_expect, "data", data)


def checking_type_read_excel_by_sheet(filepath,sheet_name,col):
    excel_data_df = pandas.read_excel(open(filepath,'rb'), sheet_name=sheet_name)

    # Apply the strip method to each cell to remove leading and trailing spaces
    excel_data_df = excel_data_df.apply(lambda col: col.map(lambda x: x.strip() if isinstance(x, str) else x))
    
    # Convert the cleaned DataFrame to JSON
    json_str = excel_data_df.to_json(orient='records')
    json_obj = json.loads(json_str)

    # Get index row
    row_count = len(excel_data_df.index)

    for i in range(row_count):
        print("INDEX_>: ",i, "DATA: ",json_obj[i][col])

    for i in range(row_count):
        print("CHECKING: ",col," INDEX: ",i, "DATA: ",json_obj[i][col]," TYPE: ",type(json_obj[i][col]))



# checking_type_read_excel_by_sheet("D:/automate/RobotFramework/File/User_Login.xlsx","User","Username")
