import json
import pandas as pd     
import openpyxl 
from openpyxl import load_workbook
import warnings

from openpyxl.utils.exceptions import InvalidFileException
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# update 13/12/2024 by Chanok

def Read_Excel_by_Sheet(filepath, sheet_name):

    # อ่านข้อมูลจาก Excel ด้วย pandas
    excel_data_df = pd.read_excel(filepath, sheet_name=sheet_name, dtype=str)

    # ใช้เมธอด strip เพื่อลบช่องว่างด้านหน้าและด้านหลังในแต่ละเซลล์
    excel_data_df = excel_data_df.apply(lambda col: col.map(
        lambda x: x.strip() if isinstance(x, str) else x))

    # แปลง DataFrame ที่ทำความสะอาดแล้วเป็น JSON
    json_str = excel_data_df.to_json(orient='records')
    json_obj = json.loads(json_str)

    # นับจำนวนแถว
    row_count = len(excel_data_df.index)
    # นับจำนวนคอลัมน์
    column_count = len(excel_data_df.columns)

   # ตรวจสอบว่า json_obj เป็นค่าว่างหรือไม่
    if not json_obj:  # ถ้า json_obj เป็น [] หรือว่าง
        json_obj = []
        row_count = 0
        column_count = 0
        return json_obj, row_count, column_count
    

    return json_obj, row_count, column_count


def Write_Excel_by_Sheet(filepath, sheet_name, row_expect, column_expect, data):
    book = openpyxl.load_workbook(filepath)
    sheet = book[sheet_name]
    sheet.cell(row=int(row_expect), column=int(column_expect)).value = data.strip()
    book.save(filepath)
    book.close()

